# 신규생성 : 2022.02.19 김재민
# 개요 : 공공API로부터 공휴일 데이터를 받아와 엑셀에 저장
#     : year, year+1을 받아서 엑셀에 연달아 데이터를 작성해야함
# 향후 개발 방향 : year를 StartSetHoliday()인수로 넣음.
# 수정 : 2022.03.29 김재민 : holiday.xlsx 파일에 주말 리스트도 같이 넣음 #001

import requests
import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pandas as pd

# 변수선언 START
mykey = 'McQ4cSLtaKSUaWUqWP9ogxYIkoakYViOMDeXF4m5TMETSh%2BuxDgHZNasTqWmH5PNOxgRtlumXp9%2BtLqZvu4vsw%3D%3D'
decordingKey = 'McQ4cSLtaKSUaWUqWP9ogxYIkoakYViOMDeXF4m5TMETSh+uxDgHZNasTqWmH5PNOxgRtlumXp9+tLqZvu4vsw=='
todayDate = datetime.datetime.now().strftime('%Y%m%d')
#year = todayDate[0:4]
fileDirPath = 'C:/Users/KJM/Desktop/DSVAN20220214/' #TestCode
holidayFileName = fileDirPath + 'holiday.xlsx' #TestCode
# 변수선언 END


def print_whichday(year, month, day):
    r = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일']
    aday = datetime.date(year, month, day)
    bday = aday.weekday()
    return r[bday]

def get_request_query(url, operation, params, serviceKey):
    import urllib.parse as urlparse
    params = urlparse.urlencode(params)
    request_query = url + '/' + operation + '?' + params + '&' + 'serviceKey' + '=' + serviceKey
    return request_query

# 일반 인증키(Encoding)

def startSetHoliday(year, filePath) :
    # input : filePath = 'C:/User/KJM/Desktop/DSVAN+todaydate'
    tempStrIndex = filePath.find('DSVAN')
    fileDirPath = filePath[:tempStrIndex]
    holidayFileName = fileDirPath + 'holiday.xlsx'
    print('holiday.xlsx 파일 경로 : ', holidayFileName)
    wb = load_workbook(holidayFileName)
    ws = wb.active
    rowIndex = ws.max_row
    if(rowIndex != 1) :
        rowIndex = rowIndex + 1
    for month in range(1, 13):

        if month < 10:
            month = '0' + str(month)
        else:
            month = str(month)

        url = 'http://apis.data.go.kr/B090041/openapi/service/SpcdeInfoService'
        # 공휴일 정보 조회
        operation = 'getRestDeInfo'
        params = {'solYear': year, 'solMonth': month}

        request_query = get_request_query(url, operation, params, mykey)
        get_data = requests.get(request_query)

        if True == get_data.ok:
            soup = BeautifulSoup(get_data.content, 'html.parser')

            item = soup.findAll('item')
            # print(item);
            for i in item:
                day = int(i.locdate.string[-2:])
                weekname = print_whichday(int(year), int(month), day)
                print(i.datename.string, i.isholiday.string, i.locdate.string, weekname)
                ws.cell(row=rowIndex, column=1, value=i.datename.string)
                ws.cell(row=rowIndex, column=2, value=i.isholiday.string)
                ws.cell(row=rowIndex, column=3, value=i.locdate.string)
                ws.cell(row=rowIndex, column=4, value=weekname)
                rowIndex = rowIndex + 1

        # # 001 START
        # todayDateYear = todayDate[0:4]
        # startDate = todayDateYear+'0101'
        # fulldt = datetime.datetime.strptime(startDate, '%Y%m%d')
        # count = 0
        # dataframe = pd.DataFrame(columns={'비고', 'check', '날짜', '요일'})
        #
        # while(count < 732) :
        #     fulldt = fulldt + datetime.timedelta(days=count)
        #     if(fulldt.weekday() == 5):
        #         dataframe = pd.DataFrame(data=[['토요일',datetime.datetime.strptime(fulldt, '%Y%m%d'), 'Y', '토요일']], columns=['비고', 'check', '날짜', '요일'])
        #
        #     if(fulldt.weekday() == 6):
        #         dataframe = pd.DataFrame(data=[['일요일',datetime.datetime.strptime(fulldt, '%Y%m%d'), 'Y', '일요일']], columns=['비고', 'check', '날짜', '요일'])
        #
        #
        #     count = count + 1



            # dataframe에 write

        # 001 END



        wb.save(holidayFileName)
        wb.close()
