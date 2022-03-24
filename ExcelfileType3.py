# 신규생성 : 2022.01.31 김재민
# 개요 : 1000JISINCHEON, 1111JISGUNSAN 파일에 대한 데이터 추출 함수
# 수정 : 2022.02.03 김재민 : checkValue 값 체크로직 삭제
#       2022.02.03 김재민 : 발주계획 엑셀 파일 write 함수 call 부분 추가 #001
#       2022.02.03 김재민 : 품번, 납기일 전역변수 추가 및 납기일 Split #002
#       2022.02.04 김재민 : startWriteCell() 함수 호출을 위한 변수선언 및 함수 호출 #003
#       2022.02.13 김재민 : 데이터가 1개일때, 2개일때 함수 call 로직 추가
#       2022.03.24 김재민 : 납품처 탐색 범위 하드코딩에서 다이나믹으로 변경 #004


import pandas as pd
import numpy as np
import WriteReleasePlan
import datetime
from openpyxl import load_workbook



# DataFrame 기본 옵션 세팅 START
pd.set_option('display.max_seq_items', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)


# DataFrame 기본 옵션 세팅 END

def getStartData(path, fileName, wbFailedListExcel, pastReleaseWorkBook, pastReleaseWorkSheet, todayDate):
    # input - path : 'C:/Users/KJM/Desktop/DSVAN'+todayDate
    # input - fileName : 1000INCHOEN.xlsx
    # input - wbFailedListExcel : load_workbook(실패한 데이터를 작성할 엑셀)

    # 변수선언 START
    itemNumber = None;  # 품번 #002
    releaseDate = None;  # 납기일 #002
    rowFr = None  # 003
    rowTo = None  # 003
    fixColumn = 3  # 003
    columnFr = 6  # 003
    columnTo = 40  # 003
    fixRow = 4  # 003
    fileDirPath = 'C:/Users/KJM/Desktop/DSVAN'+todayDate+'/'
    releaseFileName = fileDirPath + 'doosanReleasePlan' + todayDate + '.xlsx'
    orderNumber = None
    semiOrderNumber = None
    categoryDate = None
    # 변수선언 END

    pastWb = pastReleaseWorkBook
    pastWs = pastReleaseWorkSheet
    releaseWorkBook = load_workbook(fileDirPath + 'doosanReleasePlan' + todayDate + '.xlsx')
    releaseWorkSheet = releaseWorkBook.active

    # 탐색 범위 선언 START
    if '1000JISINCHEON' in fileName : #003
        print('1000JISINCHOEN 파일 시작')
        excelDataFrame = pd.read_excel(fileDirPath + '/수행예정데이터/1000JISINCHEON.xlsx',
                                       dtype={'발주번호': str,
                                              '발주항번': str})

        # 004 START
        endOfRow = len(releaseWorkSheet['B'])
        startRow = 0
        rowCount = 0

        for i in range(1, endOfRow):
            if (releaseWorkSheet.cell(i, 2).value == '인천공장' and startRow == 0):
                startRow = i
                continue

            if (releaseWorkSheet.cell(i, 2).value == '인천공장' and startRow != 0):
                rowCount = rowCount + 1
                continue
            i = i + 1

        rowFr = startRow
        rowTo = startRow + rowCount + 1
        # 004 END

        excelDataFrame.drop(excelDataFrame.columns[0], axis=1, inplace=True)
        # rowFr = 11
        # rowTo = 50

    elif '1111JISGUNSAN' in fileName : #003
        print('1111JISGUNSAN 파일 시작')
        excelDataFrame = pd.read_excel(fileDirPath + '/수행예정데이터/1111JISGUNSAN.xlsx',
                                       dtype={'발주번호': str,
                                              '발주항번': str})

        # 004 START
        endOfRow = len(releaseWorkSheet['B'])
        startRow = 0
        rowCount = 0

        for i in range(1, endOfRow):
            if (releaseWorkSheet.cell(i, 2).value == '군산공장' and startRow == 0):
                startRow = i
                continue

            if (releaseWorkSheet.cell(i, 2).value == '군산공장' and startRow != 0):
                rowCount = rowCount + 1
                continue
            i = i + 1

        rowFr = startRow
        rowTo = startRow + rowCount + 1
        # 004 END

        excelDataFrame.drop(excelDataFrame.columns[0], axis=1, inplace=True)
        # rowFr = 50
        # rowTo = 58
    else :
        print('파일 분류 에러 : ExcelfileType3')
    # 탐색 범위 선언 END

    releaseWorkBook.close()
    print('START : %s' % fileName)
    print('rowFr : %d' % rowFr)
    print('rowTo : %d' % rowTo)
    print('▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼')
    # excelDataFrame = pd.read_excel(fileName, usecols=[6, 9, 15, 29, 33, 8],
    #                                dtype={'발주번호':str,
    #                                       '발주항번':str})
    # 발주번호, 품번, Category, 납기일, 요청수량, 발주항번
    # excelDataFrame = pd.read_excel(fileName, names=['발주번호', '발주항번', '품번', 'Category', '납기일(Actual)', '요청수량'],
    #                                dtype={'발주번호': str,
    #                                       '발주항번': str})

    # 발주번호가 NULL값인 행 제거 START
    excelDataFrame['발주번호'].replace('', np.nan, inplace=True)
    excelDataFrame.dropna(subset=['발주번호'], inplace=True)
    # 발주번호가 NULL값인 행 제거 END

    # Category 컬럼이 Q이거나 R이 아닌 행 제거 START
    for index, row in excelDataFrame.iterrows():
        if (row['Category'] != 'Q' and row['Category'] != 'R'):
            excelDataFrame.drop(index, inplace=True)
        else:
            continue
    # Category 컬럼이 Q이거나 R이 아닌 행 제거 END

    # 발주번호 str로 형변환 START
    excelDataFrame['발주번호'] = excelDataFrame['발주번호'].astype(str)  # 차후 소수점 자르는 로직 필요
    excelDataFrame['발주항번'] = excelDataFrame['발주항번'].astype(str)
    # 발주번호 str로 형변환 END

    # dataframe index 재선언 START
    print('전체 인덱스 개수 : %d' % len(excelDataFrame))
    print('-----------------------------------------------------')
    newIdxArr = []
    for i in range(len(excelDataFrame)):
        newIdxArr.append(i)
    excelDataFrame.set_index(keys=[newIdxArr], inplace=True)
    print('-----------------------------------------------------')
    # dataframe index 재선언 END

    orderCount = 0
    checkValue = False

    # 데이터 추출 로직 START
    # 데이터가 1개 있을 때 실행되는 로직 START
    if (len(excelDataFrame) == 1):
        # print('데이터가 1개, 예외처리 필요')
        orderCount = excelDataFrame.iloc[0, 5]
        print('납품수량 합계 : %d' % orderCount)
        print('발주번호 : %s' % excelDataFrame.iloc[0, 0])  # 발주번호
        print('발주항번 : %s '% excelDataFrame.iloc[0, 1])
        print('품번 : %s' % excelDataFrame.iloc[0, 2])  # 품번
        print('Category : %s' % excelDataFrame.iloc[0, 3])  # Category
        print('납기일 : %s' % excelDataFrame.iloc[0, 4])  # 납기일
        print('요청수량 : %d' % excelDataFrame.iloc[0, 5])  # 요청수량 INTEGER
        print('ExcelWrite Function Call')
        categoryDate = excelDataFrame.iloc[0, 3]
        itemNumber = excelDataFrame.iloc[0, 2]
        releaseDate = excelDataFrame.iloc[0, 4]
        orderNumber = excelDataFrame.iloc[0, 0]
        semiOrderNumber = excelDataFrame.iloc[0, 1]
        WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                        fixColumn, columnFr, columnTo,
                                        fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                        , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
    # 데이터가 1개 있을 때 실행되는 로직 END

    # 데이터가 2개 있을 떄 실행되는 로직 START
    elif (len(excelDataFrame) == 2):
        if (excelDataFrame.iloc[0, 2] == excelDataFrame.iloc[1, 2]):
            if (excelDataFrame.iloc[0, 4] == excelDataFrame.iloc[1, 4]):
                # 동일품번 동일납기
                orderCount = excelDataFrame.iloc[0, 5] + excelDataFrame.iloc[1, 5]
                print('납품수량 합계 : %d' % orderCount)
                print('발주번호 : %s' % excelDataFrame.iloc[0, 0])  # 발주번호
                print('발주항번 : %s ' % excelDataFrame.iloc[0, 1])
                print('품번 : %s' % excelDataFrame.iloc[0, 2])  # 품번
                print('Category : %s' % excelDataFrame.iloc[0, 3])  # Category
                print('납기일 : %s' % excelDataFrame.iloc[0, 4])  # 납기일
                print('요청수량 : %d' % excelDataFrame.iloc[0, 5])  # 요청수량 INTEGER
                print('ExcelWrite Function Call')
                categoryDate = excelDataFrame.iloc[0, 3]
                itemNumber = excelDataFrame.iloc[0, 2]
                releaseDate = excelDataFrame.iloc[0, 4]
                orderNumber = excelDataFrame.iloc[0, 0]
                semiOrderNumber = excelDataFrame.iloc[0, 1]
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
                print('-----------------------------------------------------')
            else:
                # 동일품번 다른납기
                orderCount = excelDataFrame.iloc[0, 5]
                print('납품수량 합계 : %d' % orderCount)
                print('발주번호 : %s' % excelDataFrame.iloc[0, 0])  # 발주번호
                print('발주항번 : %s ' % excelDataFrame.iloc[0, 1])
                print('품번 : %s' % excelDataFrame.iloc[0, 2])  # 품번
                print('Category : %s' % excelDataFrame.iloc[0, 3])  # Category
                print('납기일 : %s' % excelDataFrame.iloc[0, 4])  # 납기일
                print('요청수량 : %d' % excelDataFrame.iloc[0, 5])  # 요청수량 INTEGER
                print('ExcelWrite Function Call')
                categoryDate = excelDataFrame.iloc[0, 3]
                itemNumber = excelDataFrame.iloc[0, 2]
                releaseDate = excelDataFrame.iloc[0, 4]
                orderNumber = excelDataFrame.iloc[0, 0]
                semiOrderNumber = excelDataFrame.iloc[0, 1]
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
                print('-----------------------------------------------------')
                orderCount = excelDataFrame.iloc[1, 5]
                print('납품수량 합계 : %d' % orderCount)
                print('발주번호 : %s' % excelDataFrame.iloc[1, 0])  # 발주번호
                print('발주항번 : %s ' % excelDataFrame.iloc[1, 1])
                print('품번 : %s' % excelDataFrame.iloc[1, 2])  # 품번
                print('Category : %s' % excelDataFrame.iloc[1, 3])  # Category
                print('납기일 : %s' % excelDataFrame.iloc[1, 4])  # 납기일
                print('요청수량 : %d' % excelDataFrame.iloc[1, 5])  # 요청수량 INTEGER
                print('ExcelWrite Function Call')
                categoryDate = excelDataFrame.iloc[1, 3]
                itemNumber = excelDataFrame.iloc[1, 2]
                releaseDate = excelDataFrame.iloc[1, 4]
                orderNumber = excelDataFrame.iloc[1, 0]
                semiOrderNumber = excelDataFrame.iloc[1, 1]
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
                print('-----------------------------------------------------')
        else:
            # 다른품번
            orderCount = excelDataFrame.iloc[0, 5]
            print('납품수량 합계 : %d' % orderCount)
            print('발주번호 : %s' % excelDataFrame.iloc[0, 0])  # 발주번호
            print('발주항번 : %s ' % excelDataFrame.iloc[0, 1])
            print('품번 : %s' % excelDataFrame.iloc[0, 2])  # 품번
            print('Category : %s' % excelDataFrame.iloc[0, 3])  # Category
            print('납기일 : %s' % excelDataFrame.iloc[0, 4])  # 납기일
            print('요청수량 : %d' % excelDataFrame.iloc[0, 5])  # 요청수량 INTEGER
            print('ExcelWrite Function Call')
            categoryDate = excelDataFrame.iloc[0, 3]
            itemNumber = excelDataFrame.iloc[0, 2]
            releaseDate = excelDataFrame.iloc[0, 4]
            orderNumber = excelDataFrame.iloc[0, 0]
            semiOrderNumber = excelDataFrame.iloc[0, 1]
            WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                            fixColumn, columnFr, columnTo,
                                            fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                            , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
            print('-----------------------------------------------------')
            orderCount = excelDataFrame.iloc[1, 5]
            print('납품수량 합계 : %d' % orderCount)
            print('발주번호 : %s' % excelDataFrame.iloc[1, 0])  # 발주번호
            print('발주항번 : %s ' % excelDataFrame.iloc[1, 1])
            print('품번 : %s' % excelDataFrame.iloc[1, 2])  # 품번
            print('Category : %s' % excelDataFrame.iloc[1, 3])  # Category
            print('납기일 : %s' % excelDataFrame.iloc[1, 4])  # 납기일
            print('요청수량 : %d' % excelDataFrame.iloc[1, 5])  # 요청수량 INTEGER
            print('ExcelWrite Function Call')
            categoryDate = excelDataFrame.iloc[1, 3]
            itemNumber = excelDataFrame.iloc[1, 2]
            releaseDate = excelDataFrame.iloc[1, 4]
            orderNumber = excelDataFrame.iloc[1, 0]
            semiOrderNumber = excelDataFrame.iloc[1, 1]
            WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                            fixColumn, columnFr, columnTo,
                                            fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                            , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
            print('-----------------------------------------------------')
    # 데이터가 2개 있을 때 실행되는 로직 END

    # 데이터가 3개 이상 있을 때 실행되는 로직 START
    for i in range(len(excelDataFrame) - 1):
        print('-----------------------------------------------------')
        if (excelDataFrame.iloc[i, 2] == excelDataFrame.iloc[i + 1, 2]):
            if (excelDataFrame.iloc[i, 4] == excelDataFrame.iloc[i + 1, 4]):
                if (i >= len(excelDataFrame) - 2):
                    #checkValue = True
                    orderCount = orderCount + excelDataFrame.iloc[i + 1, 5]
                    categoryDate = excelDataFrame.iloc[i, 3]
                    print('납품수량 합계 : %d' % orderCount)
                    print('발주번호 : %s' % excelDataFrame.iloc[i, 0])  # 발주번호
                    print('발주항번 : %s ' % excelDataFrame.iloc[i, 1])
                    print('품번 : %s' % excelDataFrame.iloc[i, 2])  # 품번
                    print('Category : %s' % excelDataFrame.iloc[i, 3])  # Category
                    print('납기일 : %s' % excelDataFrame.iloc[i, 4])  # 납기일
                    print('요청수량 : %d' % excelDataFrame.iloc[i, 5])  # 요청수량 INTEGER
                    print('-----------------------------------------------------')

                # 원래 동일품번 동일납기 로직 수행
                checkValue = True
                orderCount = orderCount + excelDataFrame.iloc[i + 1, 5]
                print('납품수량 합계 : %d' % orderCount)
                print('발주번호 : %s' % excelDataFrame.iloc[i+1, 0])  # 발주번호
                print('발주항번 : %s ' % excelDataFrame.iloc[i+1, 1])
                print('품번 : %s' % excelDataFrame.iloc[i+1, 2])  # 품번
                print('Category : %s' % excelDataFrame.iloc[i+1, 3])  # Category
                print('납기일 : %s' % excelDataFrame.iloc[i+1, 4])  # 납기일
                print('요청수량 : %d' % excelDataFrame.iloc[i+1, 5])  # 요청수량 INTEGER
                categoryDate = excelDataFrame.iloc[i+1, 3]
                itemNumber = excelDataFrame.iloc[i+1, 2]
                releaseDate = excelDataFrame.iloc[i+1, 4]
                orderNumber = excelDataFrame.iloc[i+1, 0]
                semiOrderNumber = excelDataFrame.iloc[i+1, 1]

                print(releaseDate)
                if (i == len(excelDataFrame) - 2):
                    print('ExcelWrite Function Call') #001
                    WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                    fixColumn, columnFr, columnTo,
                                                    fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                    , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
                print('-----------------------------------------------------')

            else:
                # 원래 동일품번 다른납기 로직 수행
                # orderCount 기록 후 0으로 초기화
                # 1 orderCount 기록
                orderCount = orderCount + excelDataFrame.iloc[i,5]
                print('납품수량 합계 : %d' % orderCount)
                print('발주번호 : %s' % excelDataFrame.iloc[i, 0])  # 발주번호
                print('발주항번 : %s ' % excelDataFrame.iloc[i, 1])
                print('품번 : %s' % excelDataFrame.iloc[i, 2])  # 품번
                print('Category : %s' % excelDataFrame.iloc[i, 3])  # Category
                print('납기일 : %s' % excelDataFrame.iloc[i, 4])  # 납기일
                print('요청수량 : %d' % excelDataFrame.iloc[i, 5])  # 요청수량 INTEGER
                print('ExcelWrite Function Call')
                categoryDate = excelDataFrame.iloc[i, 3]
                itemNumber = excelDataFrame.iloc[i, 2]
                releaseDate = excelDataFrame.iloc[i, 4]
                orderNumber = excelDataFrame.iloc[i, 0]
                semiOrderNumber = excelDataFrame.iloc[i, 1]
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
                print('-----------------------------------------------------')
                # 2 orderCount = 0 초기화
                orderCount = 0
                # 마지막 index 수행
                if (i == len(excelDataFrame) - 2):
                    print('마지막 index 실행')
                    orderCount = orderCount + excelDataFrame.iloc[i + 1, 5]
                    print('납품수량 합계 : %d' % orderCount)
                    print('발주번호 : %s' % excelDataFrame.iloc[i+1, 0])  # 발주번호
                    print('발주항번 : %s ' % excelDataFrame.iloc[i+1, 1])
                    print('품번 : %s' % excelDataFrame.iloc[i+1, 2])  # 품번
                    print('Category : %s' % excelDataFrame.iloc[i+1, 3])  # Category
                    print('납기일 : %s' % excelDataFrame.iloc[i+1, 4])  # 납기일
                    print('요청수량 : %d' % excelDataFrame.iloc[i+1, 5])  # 요청수량 INTEGER
                    print('ExcelWrite Function Call')
                    categoryDate = excelDataFrame.iloc[i+1, 3]
                    itemNumber = excelDataFrame.iloc[i+1, 2]
                    releaseDate = excelDataFrame.iloc[i+1, 4]
                    orderNumber = excelDataFrame.iloc[i + 1, 0]
                    semiOrderNumber = excelDataFrame.iloc[i + 1, 1]
                    WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                    fixColumn, columnFr, columnTo,
                                                    fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                    , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
                    print('-----------------------------------------------------')
                    orderCount = 0

        else:
            # 다른 품번
            # orderCount 기록 후 0으로 초기화
            # 1 orderCount 기록
            orderCount = orderCount + excelDataFrame.iloc[i,5]
            print('납품수량 합계 : %d' % orderCount)
            print('발주번호 : %s' % excelDataFrame.iloc[i, 0])  # 발주번호
            print('발주항번 : %s ' % excelDataFrame.iloc[i, 1])
            print('품번 : %s' % excelDataFrame.iloc[i, 2])  # 품번
            print('Category : %s' % excelDataFrame.iloc[i, 3])  # Category
            print('납기일 : %s' % excelDataFrame.iloc[i, 4])  # 납기일
            print('요청수량 : %d' % excelDataFrame.iloc[i, 5])  # 요청수량 INTEGER
            print('ExcelWrite Function Call')
            categoryDate = excelDataFrame.iloc[i, 3]
            itemNumber = excelDataFrame.iloc[i, 2]
            releaseDate = excelDataFrame.iloc[i, 4]
            orderNumber = excelDataFrame.iloc[i, 0]
            semiOrderNumber = excelDataFrame.iloc[i, 1]
            WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                            fixColumn, columnFr, columnTo,
                                            fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                            , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
            print('-----------------------------------------------------')
            # 2 orderCount = 0 초기화
            orderCount = 0
            # 마지막 index 수행
            if (i == len(excelDataFrame) - 2):
                print('마지막 index 실행')
                orderCount = orderCount + excelDataFrame.iloc[i + 1, 5]
                print('납품수량 합계 : %d' % orderCount)
                print('발주번호 : %s' % excelDataFrame.iloc[i+1, 0])  # 발주번호
                print('발주항번 : %s ' % excelDataFrame.iloc[i+1, 1])
                print('품번 : %s' % excelDataFrame.iloc[i+1, 2])  # 품번
                print('Category : %s' % excelDataFrame.iloc[i+1, 3])  # Category
                print('납기일 : %s' % excelDataFrame.iloc[i+1, 4])  # 납기일
                print('요청수량 : %d' % excelDataFrame.iloc[i+1, 5])  # 요청수량 INTEGER
                print('ExcelWrite Function Call')
                categoryDate = excelDataFrame.iloc[i+1, 3]
                itemNumber = excelDataFrame.iloc[i+1, 2]
                releaseDate = excelDataFrame.iloc[i+1, 4]
                orderNumber = excelDataFrame.iloc[i + 1, 0]
                semiOrderNumber = excelDataFrame.iloc[i + 1, 1]
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs, categoryData=categoryDate)  # 003
                print('-----------------------------------------------------')
                orderCount = 0
    # 데이터가 3개 이상 있을 때 실행되는 로직 END
    # 데이터 추출 로직 END

    print('▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲')
