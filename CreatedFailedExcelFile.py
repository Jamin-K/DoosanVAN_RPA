# 신규생성 : 2022.03.03 김재민
# 개요 : 실패 데이터를 담는 엑셀 파일생성, 1000JISINCHEON, 6000GUNSAN 등의 파일 이름 시트를 가짐

import openpyxl

def start(path, fileName) :
    wb = openpyxl.Workbook()
    newFileName = path + '/FailedData/' + fileName

    # 시트생성 START
    wb.active.title = '1000DirINCHEON'
    wb.create_sheet('1000INCHEON')
    wb.create_sheet('1100CKD')
    wb.create_sheet('1130INCHEON')
    wb.create_sheet('6000ANSAN')
    wb.create_sheet('1000JISINCHEON')
    wb.create_sheet('1111JISGUNSAN')
    # 시트생성 END

    wb.save(newFileName)
    print('파일 생성 완료!!  %s' %newFileName)