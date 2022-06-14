# 신규생성 : 2022.05.29 김재민
# 개요 : 완료데이터 엑셀 저장 후 해당 파일을 가지고 납품처가 군산공장일 경우 1일에 보내는 수량이 10%==0 이 되도록
# 묶음 처리하는 스크립트
# 당일로부터 +5DAY까지만 수행
# 수정 : yyyy.mm.dd 김재민 :


# 필요 데이터 input : 날짜, 품번, 수량
#           output : 날짜, 품번, 수량, 해당 날짜 수량 합계
# dictionary data type 사용
# 날짜를 key값으로, 품번과 수량을 value 값으료 사용
# dic의 value에는 list를 포함한 모든 자료형 가능
# if index % 2 == 0 then dic[a]의 값은 품번, else then 은 qty
# dic = {'4/1' : [itemnumber1, qty1, itemnumber2, qty2, ...]
#           , '4/2' : [itemnumber1, qty1, itemnumebr2, qty2...]
# ---> 대안 1. dic의 value 값을 리스트로 대신 함
#           (1) 엑셀 헤더에서 날짜 데이터를 추출하여 +5day 까지 key를 넣어 dic 초기화
#           (2) list자료형인 tempDDay, temp1Day, temp2Day, temp3Day, temp4Day 자료형 초기화
#           (3) 엑셀 row 를 탑다운 형태로 리스트에 append
#           (4) 해당 날짜 데이터를 리스트에 전부 append 하면 다음 날짜로 넘어가고 반복.
#           (5)
#       >> 구현 실패

# ---> 대안 2. 전체를 리스트 형태로 구현도 가능.
#           (1) temp list 변수는 동일하게 구성
#           (2) tempList = [date, 'itemNum1,qty1,itemNum2,qty2,...']
#           (3) mainList = [tempList, tempList1, tempList2, ...]
#           (4) mainList[0]의 value인 tempList[1] 을 ',' 로 split하여 결과물을 splitList 에 저장
#           (5) splitLIst의 index%2==1 인 값들을 더해서 sumQty%10 == 0 이면 mainList[0] 수행 완료
#           (6) sumQty%10 != 0 이면 mainList[1] 의 qty중 하나씩 옮겨서 %10==0을 만드는걸로


import pandas as pd
from openpyxl import load_workbook
import datetime


def Start(defaultPath, todayDate) :
    #input
    # (1) defaultPath = 'C:/Users/KJM/Desktop/DSVAN'
    # (2) todayDate  = 오늘날짜
    #Output : errorFlag
    print('--------------------------------------')
    print('군산 10개 묶음 START!!')

    tempList = []
    errorFlag = 0 # 0일때 error false, 1일때 error true
    filePath = defaultPath + todayDate + '/완료데이터/ReleasePlan.xlsx'

    releaseWorkBook = load_workbook(filePath)
    releaseWorkSheet = releaseWorkBook.active

    # row 범위 탐색 START
    endOfRow = len(releaseWorkSheet['B'])
    startRow = 0
    rowCount = 0

    for i in range(1, endOfRow):
        if (releaseWorkSheet.cell(i, 2).value == '군산공장' and startRow == 0):
            startRow = i
            continue

        if (releaseWorkSheet.cell(i, 2).value == '군산공장' and startRow != 0):
            rowCount = rowCount + 1
            continue
        i = i + 1

    rowFr = startRow
    rowTo = startRow + rowCount + 1
    # row 범위 탐색 END

    # cell 범위 탐색 START
    endOfCol = 15
    startCell = 0
    colCount = 0

    datetimeToday = datetime.datetime.strptime(todayDate, '%Y%m%d')

    tempDate = datetime.datetime.strftime(datetimeToday + datetime.timedelta(days=0), '%m/%d')
    for i in range(6, endOfCol):
        if(releaseWorkSheet.cell(4, i).value == tempDate):
            startCell=i
        else:
            colCount = colCount + 1
    colFr = startCell

    startCell = 0
    tempDate = datetime.datetime.strftime(datetimeToday + datetime.timedelta(days=5), '%m/%d')

    for i in range(colFr, endOfCol):
        if (releaseWorkSheet.cell(4, i).value == tempDate):
            startCell = i
            break
    colTo = startCell

    print('rowFr : ', rowFr)
    print('rowTo : ', rowTo)
    print('colFr : ', colFr)
    print('colTo : ', colTo)

    # cell 범위 탐색 END

    # 사용 데이터 추출 START
    tempDDay = []
    temp1Day = []
    temp2Day = []
    temp3Day = []
    temp4Day = []
    mainList = []
    tempExtractstr = ''
    for i in range(colFr, colTo):
        if(i==8): #Dday인 경우 수행
            tempDDay.append(releaseWorkSheet.cell(4, i).value)

            for j in range(rowFr, rowTo):
                if(releaseWorkSheet.cell(j, i).value is None):
                    continue
                else:
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, 3).value) + '/'
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, i).value) + '/'

            tempExtractstr = tempExtractstr[:len(tempExtractstr)-1]
            tempDDay.append(tempExtractstr) # 해당 날짜 리스트[1] 에 문자열 추가
            tempExtractstr = '' # 사용한 문자열 초기화

        elif(i==9): #D+1 수행
            temp1Day.append(releaseWorkSheet.cell(4, i).value)

            for j in range(rowFr, rowTo):
                if(releaseWorkSheet.cell(j, i).value is None):
                    continue
                else:
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, 3).value) + '/'
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, i).value) + '/'

            tempExtractstr = tempExtractstr[:len(tempExtractstr)-1]
            temp1Day.append(tempExtractstr)
            tempExtractstr = ''

        elif(i==10): #D+2 수행
            temp2Day.append(releaseWorkSheet.cell(4, i).value)

            for j in range(rowFr, rowTo):
                if (releaseWorkSheet.cell(j, i).value is None):
                    continue
                else:
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, 3).value) + '/'
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, i).value) + '/'

            tempExtractstr = tempExtractstr[:len(tempExtractstr) - 1]
            temp2Day.append(tempExtractstr)
            tempExtractstr = ''

        elif(i==11): #D+3 수행
            temp3Day.append(releaseWorkSheet.cell(4, i).value)

            for j in range(rowFr, rowTo):
                if (releaseWorkSheet.cell(j, i).value is None):
                    continue
                else:
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, 3).value) + '/'
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, i).value) + '/'

            tempExtractstr = tempExtractstr[:len(tempExtractstr) - 1]
            temp3Day.append(tempExtractstr)
            tempExtractstr = ''

        elif(i==12): #D+4 수행
            temp4Day.append(releaseWorkSheet.cell(4, i).value)

            for j in range(rowFr, rowTo):
                if (releaseWorkSheet.cell(j, i).value is None):
                    continue
                else:
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, 3).value) + '/'
                    tempExtractstr = tempExtractstr + str(releaseWorkSheet.cell(j, i).value) + '/'

            tempExtractstr = tempExtractstr[:len(tempExtractstr) - 1]
            temp4Day.append(tempExtractstr)
            tempExtractstr = ''


    print(tempDDay)
    print(temp1Day)
    print(temp2Day)
    print(temp3Day)
    print(temp4Day)


    for i in range(rowFr, rowTo):
        if(releaseWorkSheet.cell(i,colFr).value is None) :
            continue
        else:
            tempDDay.append(releaseWorkSheet.cell(i, 3).value)
            tempDDay.append(releaseWorkSheet.cell(i, colFr).value)

    for i in range(rowFr, rowTo):
        if(releaseWorkSheet.cell(i,colFr+1).value is None) :
            continue
        else:
            temp1Day.append(releaseWorkSheet.cell(i, 3).value)
            temp1Day.append(releaseWorkSheet.cell(i, colFr+1).value)

    for i in range(rowFr, rowTo):
        if(releaseWorkSheet.cell(i,colFr+2).value is None) :
            continue
        else:
            temp2Day.append(releaseWorkSheet.cell(i, 3).value)
            temp2Day.append(releaseWorkSheet.cell(i, colFr+2).value)

    for i in range(rowFr, rowTo):
        if(releaseWorkSheet.cell(i,colFr+3).value is None) :
            continue
        else:
            temp3Day.append(releaseWorkSheet.cell(i, 3).value)
            temp3Day.append(releaseWorkSheet.cell(i, colFr+3).value)

    for i in range(rowFr, rowTo):
        if(releaseWorkSheet.cell(i,colFr+4).value is None) :
            continue
        else:
            temp4Day.append(releaseWorkSheet.cell(i, 3).value)
            temp4Day.append(releaseWorkSheet.cell(i, colFr+4).value)



    # 사용 데이터 추출 END















