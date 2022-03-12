# 신규생성 : 2022.02.03 김재민
# 개요 : doosanReleasePlan 엑셀 세팅
# 수정 : 2022.02.04 김재민 : 함수 input값 추가
#       2022.02.19 김재민 : 클래스 생성, 함수 호출 및 getter 접근 #001
#       2022.02.19 김재민 : 얻은 좌표에 Write OrderCount 및 엑셀 저장 #002
#       2022.03.02 김재민 : ReleasePlan에 쓰기 실패한 데이터 리스트 기록 in writeFaieldlist.xlsx #003
#       2022.03.12 김재민 : D-1 완료 데이터와 합쳐서 D-day ReleasePlan에 작성 #004



from openpyxl import load_workbook
from findValueLocation import Coordinate
import CheckWorkDays
import datetime
import numpy as np


def startWriteCell(filePath, rowFr, rowTo, fixColumn, columnFr, columnTo, fixRow, findValue1, findValue2, orderCount,
                   orderNumber, semiOrderNumber, wbFailedListExcel, fileName,
                   jisData=None, categoryData=None) :
    # findValue1 = 품명
    # findValue2 = 날짜

    # D-1 날짜 구하기 START #004
    #tempDate = datetime.datetime.now().strftime('%Y%m%d')
    tempDate = datetime.datetime.now() + datetime.timedelta(days=-1)
    oneDaysAgoDate = datetime.datetime.strftime(tempDate, '%Y%m%d')
    pastfilePath = 'C:/Users/KJM/Desktop/DSVAN20220213/완료데이터/' + 'ReleasePlan.xlsx' #testCode
    # D-1 날짜 구하기 END

    # FailedFileName 추출 START #003
    findIndex = fileName.find('doosan') + 6
    sheetName = fileName[findIndex:]
    sheetName = sheetName[0:-13]
    sheetName = sheetName.strip()
    print('sheeteName!! %s' % sheetName)
    # FailedFileName 추출 END

    #wb = load_workbook(filePath)
    wb = load_workbook('C:/users/KJM/Desktop/DSVAN20220214/완료데이터/ReleasePlan.xlsx') #testCode
    ws = wb.active
    cordinate = Coordinate() #001
    itemNumber = findValue1

    # 공휴일 및 주말 체크 함수 START
    findValue2 = CheckWorkDays.checkHolidays(findValue2)
    # 공휴일 및 주말 체크 함수 END

    releaseDate = findValue2 #yyyy/mm/dd 형태로 받아옴

    cordinate.startFindValue(filePath, rowFr, rowTo, fixColumn, columnFr, columnTo, fixRow, itemNumber, releaseDate, ws) #001
    if(cordinate.getRow() != 0 and cordinate.getCol() != 0) : #001
        print('Write Cordinate!! : %d %d' % (cordinate.getRow(), cordinate.getCol())) #001
        print('Write Cordinate of OrderCount : %d' % orderCount) #001
        #ws.cell(cordinate.getRow(), cordinate.getCol(), orderCount) #002

        # D-1 ReleaseData 파일 열기 START #004
        pastWb = load_workbook(pastfilePath)
        pastWs = pastWb.active
        if (pastWs.cell(cordinate.getRow(), cordinate.getCol() + 1).value == None):
            ws.cell(cordinate.getRow(), cordinate.getCol(), orderCount)  # 002
        else:
            testvalue1 = pastWs.cell(cordinate.getRow(), cordinate.getCol() + 1).value + orderCount
            print('더한 데이터 : %d' % testvalue1)
            ws.cell(cordinate.getRow(), cordinate.getCol(),
                    orderCount + pastWs.cell(cordinate.getRow(), cordinate.getCol() + 1).value)

        pastWb.close()

        # D-1 ReleaseData 파일 열기 END

    elif(cordinate.getRow() == 0 and cordinate.getCol() == 0) : #003
        print('Write Failed 데이터 엑셀 기록')

        # sheetName으로 해당 시트에 기록 START
        #wsFailedListExcel = wbFailedListExcel.get_sheet_by_name('sheetName')
        wsFailedListExcel = wbFailedListExcel[sheetName]
        wsFailedListExcel.cell(1, 1, '발주번호')
        wsFailedListExcel.cell(1, 2, '발주항번')
        wsFailedListExcel.cell(1, 3, '품명')
        wsFailedListExcel.cell(1, 4, '날짜')
        wsFailedListExcel.cell(1, 5, '발주수량')
        wsFailedListExcel.cell(1, 6, 'JIS')
        wsFailedListExcel.cell(1, 7, 'Category')
        tempList = []
        tempList.append(orderNumber)
        tempList.append(semiOrderNumber)
        tempList.append(findValue1)
        tempList.append(findValue2)
        tempList.append(orderCount)
        tempList.append(jisData)
        tempList.append(categoryData)
        wsFailedListExcel.append(tempList)

        # sheetName으로 해당 시트에 기록 END


    #wb.save(filePath) #002
    wb.save('C:/users/KJM/Desktop/DSVAN20220214/완료데이터/ReleasePlan.xlsx') #testCode
    wb.close()


