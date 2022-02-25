# 신규생성 : 2022.02.04 김재민
# 개요 : 가로 세로 2개의 값을 입력받아 해당 셀의 위치를 반환하는 함수
# 수정 : 2022.02.14 김재민 : 엑셀시트에 해당 품명이 존재 여부를 체크하는 로직 추가 #001
#       2022.02.14 김재민 : 좌표 클래스 생성 #002
#       2022.02.19 김재민 : startFindValue 함수 인자 self 추가 #003
#       2022.02.19 김재민 : yyyy/mm/dd 날짜를 저장하기 위한 변수 선언(공휴일 체크에 사용) 추가 #004

from openpyxl import load_workbook

# 변수선언 START
checkFindValue = False #001
checkFindValueItem = False #001
fullDate = None #004

# 변수선언 END

class Coordinate : #002
    row = 0
    column = 0
    def __init__(self):
        self.row = 0
        self.column = 0

    def setRow(self, row):
        self.row = row

    def setCol(self, column):
        self.column = column

    def getRow(self):
        return self.row

    def getCol(self):
        return self.column

    def startFindValue(self, filePath, rowFr, rowTo, fixColumn, columnFr, columnTo, fixRow, findValue1, findValue2, ws): #003
        # ws를 한번 더 호출하지 않고 함수 인자에 담아 들고오기 때문에 엑셀을 한번 더 열 필요가 없어져 수행시간 빨라짐
        #wb = load_workbook(filePath)
        #ws = wb.active

        # findvalue2는 날짜, yyyy/mm/dd 형태로 받아옴, mm/dd 형태로 가공
        fullDate = findValue2 #004
        findValue2 = findValue2[5:10]

        global checkFindValue
        global checkFindValueItem

        for rowIndex in range(rowFr, rowTo):
            rowTempData = ws.cell(row=rowIndex, column=fixColumn).value
            if rowTempData == findValue1:
                checkFindValueItem = True  # 001

                for columnIndex in range(columnFr, columnTo):
                    columnTempData = ws.cell(row=fixRow, column=columnIndex).value
                    if columnTempData == findValue2:
                        checkFindValue = True  # 001
                        self.setRow(rowIndex) #003
                        self.setCol(columnIndex) #003
                        print('Success Find Value!! %d %d' % (self.getRow(), self.getCol())) #003
                        break
                    else:
                        checkFindValue = False  # 001
                break
            else:
                checkFindValueItem = False  # 001
        if (checkFindValue == False and checkFindValueItem == False):  # 001
            print('Failed Find Value!!(품목없음)%s %s' % (findValue1, findValue2))

        elif (checkFindValue == False and checkFindValueItem == True):  # 001
            print('Failed Find Value!!(날짜경과)%s %s' % (findValue1, findValue2))

        checkFindValue = False  # 001
        checkFindValueItem = False  # 001

        #wb.close()


