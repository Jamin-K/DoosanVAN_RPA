# 신규생성 : 2022.02.03 김재민
# 개요 : doosanReleasePlan 엑셀 파일 초기 세팅
# 수정 : 2022.03.29 김재민 : 세팅이 끝난 doosanReleasePlan+todayDate를 완료데이터/ReleasePlan.xlsx 로 복사 및 이름변경 #001
#       2022.03.30 김재민 : D-1 ReleaseData를 왼쪽으로 한 칸씩 당겨 D-day ReleaseData에 삽입 #002

from openpyxl import load_workbook
from datetime import datetime, timedelta
import shutil

def setStartPlanFile(planFilePath, todayDate, oneDaysAgoDate,defaultPath) :
    # input - defaultPath : 'C:/Users/KJM/Desktop/DSVAN'

    shutil.copyfile(planFilePath, defaultPath+todayDate+'/doosanReleasePlan' + todayDate + '.xlsx')

    print('SetExcel START!!')
    # today -2 ~ today + 32 날짜 세팅 START
    wb = load_workbook(defaultPath+todayDate+'/doosanReleasePlan' + todayDate + '.xlsx')
    ws = wb.active
    rows = 4
    columns = 6
    for i in range(-2,32) :
        date = datetime.strptime(todayDate, '%Y%m%d') + timedelta(days=i)
        #date = datetime.now() + timedelta(days=i)
        date = date.strftime('%m') + '/' + date.strftime('%d')
        ws.cell(row=rows, column=columns, value=date)
        columns = columns + 1
    wb.save(defaultPath+todayDate+'/doosanReleasePlan' + todayDate + '.xlsx')
    wb.close()
    # today -2 ~ today + 32 날짜 세팅 END

    # 001 START
    shutil.copyfile(defaultPath+todayDate+'/doosanReleasePlan' + todayDate + '.xlsx', defaultPath+todayDate+'/완료데이터/ReleasePlan.xlsx')
    # 001 END

    # 002 START
    pastWb = load_workbook(defaultPath + oneDaysAgoDate + '/완료데이터/ReleasePlan.xlsx')
    pastWs = pastWb.active

    presentWb = load_workbook(defaultPath + todayDate + '/완료데이터/ReleasePlan.xlsx')
    presentWs = presentWb.active

    pastMaxRow = pastWs.max_row
    presentMaxRow = presentWs.max_row
    tempList = list()

    for pastRowIndex in range(5, pastMaxRow):
        customer = pastWs.cell(pastRowIndex, 2).value
        itemNum = pastWs.cell(pastRowIndex, 3).value

        for pastColIndex in range(7, 40) :
            if(pastWs.cell(pastRowIndex, pastColIndex).value == None):
                tempList.append('')
            else:
                tempList.append(pastWs.cell(pastRowIndex, pastColIndex).value)

        for presentRowIndex in range(5, presentMaxRow):
            if(presentWs.cell(presentRowIndex, 2).value == customer and presentWs.cell(presentRowIndex, 3).value == itemNum):
                presentColIndex = 6
                for item in tempList:
                    presentWs.cell(row=presentRowIndex, column=presentColIndex, value=item)
                    presentColIndex = presentColIndex + 1
                tempList.clear()
                break

    pastWb.close()
    presentWb.save(defaultPath + todayDate + '/완료데이터/ReleasePlan.xlsx')
    presentWb.close()
    print('setExcel 종료!!')

    # 002 END










