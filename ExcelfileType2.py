# 신규생성 : 2022.01.31 김재민
# 개요 : 6000ANSAN 파일에 대한 데이터 추출 함수
# 수정 : 2022.02.03 김재민 : checkValue 값 체크로직 삭제
#       2022.02.03 김재민 : 발주계획 엑셀 파일 write 함수 call 부분 추가 #001
#       2022.02.03 김재민 : 품번, 납기일 전역변수 추가 및 납기일 Split #002
#       2022.02.04 김재민 : startWriteCell() 함수 호출을 위한 변수선언 및 함수 호출 #003
#       2022.02.13 김재민 : 데이터가 1개일떄, 2개일때 함수 call 로직 추가
#       2022.03.24 김재민 : 납품처 탐색 범위 하드코딩에서 다이나믹으로 변경 #004

import pandas as pd
import WriteReleasePlan
from openpyxl import load_workbook



# DataFrame 기본 옵션 세팅 START
pd.set_option('display.max_seq_items', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
# DataFrame 기본 옵션 세팅 END

def getStartData(path, fileName, wbFailedListExcel, pastReleaseWorkBook, pastReleaseWorkSheet, todayDate) :
    # input - path : 'C:/Users/KJM/Desktop/DSVAN'+todayDate
    # input - fileName : 1000INCHOEN.xlsx
    # input - wbFailedListExcel : load_workbook(실패한 데이터를 작성할 엑셀)

    # 변수선언 START
    itemNumber = None  # 품번 #002
    releaseDate = None  # 납기일 #002
    rowFr = None
    rowTo = None
    fixColumn = 3
    columnFr = 6
    columnTo = 40
    fixRow = 4
    fileDirPath = 'C:/Users/KJM/Desktop/DSVAN'+todayDate+'/'
    releaseFileName = fileDirPath + 'doosanReleasePlan' + todayDate + '.xlsx'
    # fileDirPath = 'C:/Users/KJM/Desktop/DSVAN20220214/'  # TestCode
    # releaseFileName = fileDirPath + 'doosanReleasePlan20220214.xlsx'  # TestCode
    orderNumber = None
    semiOrderNumber = None
    # 변수선언 END

    pastWb = pastReleaseWorkBook
    pastWs = pastReleaseWorkSheet
    releaseWorkBook = load_workbook(fileDirPath + 'doosanReleasePlan' + todayDate + '.xlsx')
    releaseWorkSheet = releaseWorkBook.active

    if '6000ANSAN' in fileName :
        print('6000ANSAN 파일 시작')
        excelDataFrame = pd.read_excel(fileDirPath + '/수행예정데이터/6000ANSAN.xlsx',
                                       dtype={'발주번호': str,
                                              '발주항번': str})

        # 004 START
        endOfRow = len(releaseWorkSheet['B'])
        gunsanRowCount = 0 # A/S 항목의 뚜렷한 식별자가 없기 때문에 그 전 발주처인 군산공장을 기준으로 계산
        gunsanEndRow = 0
        gunsanStartRow = 0

        for i in range(1, endOfRow) :
            if(releaseWorkSheet.cell(i, 2).value == '군산공장' and gunsanStartRow == 0) :
                gunsanStartRow = i
                continue

            if(releaseWorkSheet.cell(i, 2).value == '군산공장' and gunsanStartRow != 0) :
                gunsanRowCount = gunsanRowCount + 1
                continue
            i = i + 1

        gunsanEndRow = gunsanStartRow + gunsanRowCount + 1
        rowFr = gunsanEndRow + 3
        rowTo = endOfRow

        # 004 END
        excelDataFrame.drop(excelDataFrame.columns[0], axis=1, inplace=True)
        # rowFr = 61
        # rowTo = 133
    else :
        print('파일 분류 에러 : ExcelfileType2')

    releaseWorkBook.close()
    print('START : %s' %fileName)
    print('rowFr : %d' % rowFr)
    print('rowTo : %d' % rowTo)
    print('▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼')
    # excelDataFrame = pd.read_excel(fileName, usecols=[4, 9, 12, 45, 46],
    #                                dtype={'발주번호':str,
    #                                       '발주항번':str}) # 품번, 납품잔량, 납기일자, 발주번호, 발주항번

    # dataframe index 재선언 START
    print('전체 인덱스 개수 : %d' % len(excelDataFrame))
    print('-----------------------------------------------------')
    newIdxArr = []
    for i in range(len(excelDataFrame)):
        newIdxArr.append((i))
    excelDataFrame.set_index(keys=[newIdxArr], inplace=True)
    # dataframe index 재선언 END

    orderCount = 0
    checkValue = False

    # 데이터 추출 로직 START
    # 데이터가 1개일 때 실행되는 로직 START
    if (len(excelDataFrame) == 1) :
        orderCount = excelDataFrame.iloc[0, 1]
        print('발주번호 : %s' % excelDataFrame.iloc[0, 3])
        print('발주항번 : %s' % excelDataFrame.iloc[0, 4])
        print('납품잔량 합계 : %d' % orderCount)
        print('품번 : %s' % excelDataFrame.iloc[0, 0])
        print('납기일자 : %s' % excelDataFrame.iloc[0, 2])
        print('ExcelWrite Function Call')
        itemNumber = excelDataFrame.iloc[0, 0]
        releaseDate = excelDataFrame.iloc[0, 2]
        orderNumber = excelDataFrame.iloc[0, 3]
        semiOrderNumber = excelDataFrame.iloc[0, 4]
        WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                        fixColumn, columnFr, columnTo,
                                        fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                        , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003

    # 데이터가 1개일 때 실행되는 로직 END

    # 데이터가 2개일 때 실행되는 로직 START
    elif (len(excelDataFrame) == 2):
        if (excelDataFrame.iloc[0, 0] == excelDataFrame.iloc[1, 0]):
            if (excelDataFrame.iloc[0, 2] == excelDataFrame.iloc[1, 2]):
                # 동일품번 동일납기
                orderCount = excelDataFrame.iloc[0, 1] + excelDataFrame.iloc[1, 1]
                print('발주번호 : %s' % excelDataFrame.iloc[0, 3])
                print('발주항번 : %s' % excelDataFrame.iloc[0, 4])
                print('납품수량 합계 : %d' % orderCount)
                print('품번 : %s' % excelDataFrame.iloc[0, 0])
                print('납기일자 : %s' % excelDataFrame.iloc[0, 2])
                print('ExcelWrite Function Call')
                itemNumber = excelDataFrame.iloc[0, 0]
                releaseDate = excelDataFrame.iloc[0, 2]
                orderNumber = excelDataFrame.iloc[0, 3]
                semiOrderNumber = excelDataFrame.iloc[0, 4]
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003  # 003

            else:
                # 동일품번 다른납기
                orderCount = excelDataFrame.iloc[0, 1]
                print('발주번호 : %s' % excelDataFrame.iloc[0, 3])
                print('발주항번 : %s' % excelDataFrame.iloc[0, 4])
                print('납품잔량 합계 : %d' % orderCount)
                print('품번 : %s' % excelDataFrame.iloc[0, 0])
                print('납기일자 : %s' % excelDataFrame.iloc[0, 2])
                print('ExcelWrite Function Call')
                itemNumber = excelDataFrame.iloc[0, 0]
                releaseDate = excelDataFrame.iloc[0, 2]
                orderNumber = excelDataFrame.iloc[0, 3]
                semiOrderNumber = excelDataFrame.iloc[0, 4]
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003
                print('-----------------------------------------------------')
                orderCount = excelDataFrame.iloc[1, 1]
                print('발주번호 : %s' % excelDataFrame.iloc[0, 3])
                print('발주항번 : %s' % excelDataFrame.iloc[0, 4])
                print('납품잔량 합계 : %d' % orderCount)
                print('품번 : %s' % excelDataFrame.iloc[1, 0])
                print('납기일자 : %s' % excelDataFrame.iloc[1, 2])
                itemNumber = excelDataFrame.iloc[1, 0]
                releaseDate = excelDataFrame.iloc[1, 2]
                orderNumber = excelDataFrame.iloc[1, 3]
                semiOrderNumber = excelDataFrame.iloc[1, 4]
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003
                print('ExcelWrite Function Call')
        else :
            # 다른품번
            orderCount = excelDataFrame.iloc[0, 1]
            print('발주번호 : %s' % excelDataFrame.iloc[0, 3])
            print('발주항번 : %s' % excelDataFrame.iloc[0, 4])
            print('납품잔량 합계 : %d' % orderCount)
            print('품번 : %s' % excelDataFrame.iloc[0, 0])
            print('납기일자 : %s' % excelDataFrame.iloc[0, 2])
            print('ExcelWrite Function Call')
            itemNumber = excelDataFrame.iloc[0, 0]
            releaseDate = excelDataFrame.iloc[0, 2]
            orderNumber = excelDataFrame.iloc[0, 3]
            semiOrderNumber = excelDataFrame.iloc[0, 4]
            WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                            fixColumn, columnFr, columnTo,
                                            fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                            , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003
            print('-----------------------------------------------------')
            orderCount = excelDataFrame.iloc[1,1]
            print('발주번호 : %s' % excelDataFrame.iloc[1, 3])
            print('발주항번 : %s' % excelDataFrame.iloc[1, 4])
            print('납품잔량 합계 : %d' % orderCount)
            print('품번 : %s' % excelDataFrame.iloc[1, 0])
            print('납기일자 : %s' % excelDataFrame.iloc[1, 2])
            print('ExcelWrite Function Call')
            itemNumber = excelDataFrame.iloc[1, 0]
            releaseDate = excelDataFrame.iloc[1, 2]
            orderNumber = excelDataFrame.iloc[1, 3]
            semiOrderNumber = excelDataFrame.iloc[1, 4]
            WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                            fixColumn, columnFr, columnTo,
                                            fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                            , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003
    # 데이터가 2개일 때 실행되는 로직 END

    # 데이터가 3개 이상일 때 실행되는 로직 START
    else:
        for i in range(len(excelDataFrame) - 1):
            print('-----------------------------------------------------')
            if (excelDataFrame.iloc[i, 0] == excelDataFrame.iloc[i + 1, 0]):
                if (excelDataFrame.iloc[i, 2] == excelDataFrame.iloc[i + 1, 2]):
                    if (i >= len(excelDataFrame) - 2):
                        #checkValue = True
                        orderCount = orderCount + excelDataFrame.iloc[i + 1, 1]
                        print('발주번호 : %s' % excelDataFrame.iloc[i, 3])
                        print('발주항번 : %s' % excelDataFrame.iloc[i, 4])
                        print('납품수량 합계 : %d' % orderCount)
                        print('품번 : %s' % excelDataFrame.iloc[i, 0])
                        print('납품잔량 : %d' % orderCount)
                        print('납기일자 : %s' % excelDataFrame.iloc[i, 2])
                        print('-----------------------------------------------------')

                    # 원래 동일품번 동일납기 로직 수행
                    #checkValue = True
                    orderCount = orderCount + excelDataFrame.iloc[i + 1, 1]
                    print('발주번호 : %s' % excelDataFrame.iloc[i + 1, 3])
                    print('발주항번 : %s' % excelDataFrame.iloc[i + 1, 4])
                    print('납품수량 합계 : %d' % orderCount)
                    print('품번 : %s' % excelDataFrame.iloc[i + 1, 0])
                    print('납품잔량 : %d' % excelDataFrame.iloc[i + 1, 0])
                    print('납기일자 : %s' % excelDataFrame.iloc[i + 1, 2])
                    itemNumber = excelDataFrame.iloc[i+1, 0] #002
                    releaseDate = excelDataFrame.iloc[i+1, 2] #002
                    orderNumber = excelDataFrame.iloc[i + 1, 3]
                    semiOrderNumber = excelDataFrame.iloc[i + 1, 4]
                    if (i == len(excelDataFrame) - 2):
                        print('ExcelWrite Function Call') #001
                        WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                        fixColumn, columnFr, columnTo,
                                                        fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                        , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003
                    print('-----------------------------------------------------')

                else:
                    # 원래 동일품번 다른납기 로직 수행
                    # orderCount 기록 후 0으로 초기화
                    # 1 orderCount 기록
                    orderCount = orderCount + excelDataFrame.iloc[i, 1]
                    print('발주번호 : %s' % excelDataFrame.iloc[i, 3])
                    print('발주항번 : %s' % excelDataFrame.iloc[i, 4])
                    print('납품수량 합계 : %d' % orderCount)
                    print('품번 : %s' % excelDataFrame.iloc[i, 0])
                    print('납품잔량 : %d' % excelDataFrame.iloc[i, 1])
                    print('납기일자 : %s' % excelDataFrame.iloc[i, 2])
                    itemNumber = excelDataFrame.iloc[i, 0]  # 002
                    releaseDate = excelDataFrame.iloc[i, 2]  # 002
                    orderNumber = excelDataFrame.iloc[i, 3]
                    semiOrderNumber = excelDataFrame.iloc[i, 4]
                    print('ExcelWrite Function Call') #001
                    WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                    fixColumn, columnFr, columnTo,
                                                    fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                    , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003
                    print('-----------------------------------------------------')
                    # 2 orderCount = 0 초기화
                    orderCount = 0
                    # 마지막 index 수행
                    if (i == len(excelDataFrame) - 2):
                        print('마지막 index 실행')
                        orderCount = orderCount + excelDataFrame.iloc[i + 1, 1]
                        print('발주번호 : %s' % excelDataFrame.iloc[i+1, 3])
                        print('발주항번 : %s' % excelDataFrame.iloc[i+1, 4])
                        print('납품수량 합계 : %d' % orderCount)
                        print('품번 : %s' % excelDataFrame.iloc[i + 1, 0])
                        print('납품잔량 : %d' % excelDataFrame.iloc[i + 1, 1])
                        print('납기일자 : %s' % excelDataFrame.iloc[i + 1, 2])
                        itemNumber = excelDataFrame.iloc[i+1, 0] #002
                        releaseDate = excelDataFrame.iloc[i + 1, 2]  # 002
                        orderNumber = excelDataFrame.iloc[i+1, 3]
                        semiOrderNumber = excelDataFrame.iloc[i+1, 4]
                        # releaseDate = excelDataFrame.iloc[i + 1, 2][5:10]  # 002
                        print('ExcelWrite Function Call') #001
                        WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                        fixColumn, columnFr, columnTo,
                                                        fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                        , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003

                        orderCount = 0

            else:
                # 다른 품번
                # orderCount 기록 후 0으로 초기화
                # 1 orderCount 기록
                orderCount = orderCount + excelDataFrame.iloc[i, 1]
                print('발주번호 : %s' % excelDataFrame.iloc[i, 3])
                print('발주항번 : %s' % excelDataFrame.iloc[i, 4])
                print('납품수량 합계 : %d' % orderCount)
                print('품번 : %s' % excelDataFrame.iloc[i, 0])
                print('납품잔량 : %d' % excelDataFrame.iloc[i, 1])
                print('납기일자 : %s' % excelDataFrame.iloc[i, 2])
                itemNumber = excelDataFrame.iloc[i, 0]  # 002
                releaseDate = excelDataFrame.iloc[i, 2]  # 002
                orderNumber = excelDataFrame.iloc[i, 3]
                semiOrderNumber = excelDataFrame.iloc[i, 4]
                print('ExcelWrite Function Call')  # 001
                WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                fixColumn, columnFr, columnTo,
                                                fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003

                print('-----------------------------------------------------')
                # 2 orderCount = 0 초기화
                orderCount = 0
                # 마지막 index 수행
                if (i == len(excelDataFrame) - 2):
                    print('마지막 index 실행')
                    orderCount = orderCount + excelDataFrame.iloc[i + 1, 1]
                    print('발주번호 : %s' % excelDataFrame.iloc[i+1, 3])
                    print('발주항번 : %s' % excelDataFrame.iloc[i+1, 4])
                    print('납품수량 합계 : %d' % orderCount)
                    print('품번 : %s' % excelDataFrame.iloc[i + 1, 0])
                    print('납품잔량 : %d' % excelDataFrame.iloc[i + 1, 1])
                    print('납기일자 : %s' % excelDataFrame.iloc[i + 1, 2])
                    itemNumber = excelDataFrame.iloc[i + 1, 0]  # 002
                    releaseDate = excelDataFrame.iloc[i + 1, 2]  # 002
                    orderNumber = excelDataFrame.iloc[i + 1, 3]
                    semiOrderNumber = excelDataFrame.iloc[i + 1, 4]
                    print('ExcelWrite Function Call')  # 001
                    WriteReleasePlan.startWriteCell(releaseFileName, rowFr, rowTo,
                                                    fixColumn, columnFr, columnTo,
                                                    fixRow, itemNumber, releaseDate, orderCount, orderNumber
                                                    , semiOrderNumber, wbFailedListExcel, fileName, pastWb, pastWs)  # 003

                    print('-----------------------------------------------------')
                    orderCount = 0
    # 데이터가 3개 이상일 때 실행되는 로직 END
    # 데이터 추출 로직 END

    print('▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲')